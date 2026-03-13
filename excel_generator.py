"""
Excel Generator for Battery Pack MES
Generates Excel files from database in EXACT same format as before
CRITICAL: Format and data mapping must remain identical
"""

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font
from pathlib import Path
from datetime import datetime
import logging
from typing import Optional
import io
from database import get_qc_checks, get_all_battery_packs, get_battery_pack_info

# Standard font to override Wingdings in template (columns L/M use Wingdings which breaks WPS Office)
STANDARD_FONT = Font(name='Arial', size=10)

logger = logging.getLogger(__name__)

# CRITICAL: Use a dedicated template file that is NEVER overwritten
# template.xlsx = clean original template (read-only, never modified)
# sample.xlsx = master output file (regenerated from template + DB data)
TEMPLATE_PATH = Path(__file__).parent / "template.xlsx"
MASTER_OUTPUT_PATH = Path(__file__).parent / "sample.xlsx"

# Row mapping for Excel sheet - Duplicated here to avoid circular import
PROCESS_ROW_MAPPING = {
    "Cell Sorting": {"start_row": 8, "type": "standard"},
    "Module Assembly": {"start_row": 12, "type": "standard"},
    "Encapsulation & Soldering - Phase 1": {"start_row": 29, "type": "standard"},
    "Wire Bonding": {"start_row": 34, "type": "standard"},
    "Encapsulation Phase II (100%)": {"start_row": 35, "type": "standard"},
    "Module QC Checks": {"start_row": 37, "type": "standard"},
    "EOL Testing": {"start_row": 38, "type": "standard"},
    "Pack Assembly": {"start_row": 40, "type": "pack"},
    "Ready for Dispatch": {"start_row": 62, "type": "dispatch"}
}

# Canonical check order per process — ensures correct Excel row offsets
# regardless of the order checks were saved to DB (which is by created_at).
QC_CHECKS_ORDER = {
    "Cell Sorting": [
        "1.1.1 Acceptable range Voltage : 3.58 to 3.6 V, IR : 11.5 to 12.5 mΩ",
        "1.2.1 Jig setup on the flat surface",
        "1.2.2 Visual & spacing inspection",
        "1.2.3 No moisture, dust on the cell surface (no contaminations) - IP clean",
    ],
    "Module Assembly": [
        "2.1.1 Proper fitment to the Base plate fixture and base plate orientation",
        "2.1.2 Clinch/all rivets perpendicular to the plate & proper clinch",
        "2.2.1 Volume consistency",
        "2.2.2 Coverage in space/gap",
        "2.2.3 Jig alignment correction",
        "2.2.4 Cure and bond integrity - 2 hrs. adhesive curing after cell flip",
        "2.3.1 Cell continuity check & visual check on cell fitment, Cell Height (72.6 mm) from base plate after dismantling fixture",
        "2.3.2 Cleanliness inspection",
        "2.4.1 Even fix of sidewall (PC sheets) on all four sides without gaps, proper dimension 72.5 mm × 239.4 mm",
        "2.5.1 Placement as per drawing & wire routings (Thermistor - 7 nos / Voltage harness - 14 nos)",
        "2.5.2 Secure adhesion & positioning with thermal adhesive",
        "2.5.3 No insulation damage or routing faults (Visual + routing diagram comparison)",
        "2.6.1 Busbar fitment as per drawings, Curing time 1 hr.",
        "2.6.2 Busbar placement & defect absence",
        "2.6.3 Adhesive coverage - Ensure full bond",
        "2.6.4 Adhesion strength",
        "2.6.5 Busbar and CPT sub-assembly dim (Ht 22.2 mm from the base)",
    ],
    "Encapsulation & Soldering - Phase 1": [
        "3.1.1 No encapsulant leakage from sidewalls. Curing time 2 hr @ room temperature ~25°C",
        "3.1.2 Dimensions: L 410.3 × W 239.4 mm",
        "3.2.1 Busbar + CPT module assembly dim (Ht 78 mm from the base)",
        "3.2.2 Soldering check & Voltage of module level (soft soldering of voltage taps on cell positive surface)",
        "3.3.1 Proper Stripping, crimping & connector insertion for voltage tap and thermistor",
    ],
    "Wire Bonding": [
        "4.1.1 The 4 bonds on each cell as per the drawing, module voltage & continuity (no continuity present)",
    ],
    "Encapsulation Phase II (100%)": [
        "5.1.1 Equal distribution of encapsulant on the module (4 hrs. curing)",
        "5.2.1 Dimensions: H 78.6 × L 410.3 × W 239.4 mm",
    ],
    "Module QC Checks": [
        "6.1.1 Check for Spec: a) Weight 17 Kg  b) Module Voltage 50 V  c) Continuity: Not present  d) Module Height: 78.6 mm. Encapsulation & Final module visual checks, Floating Voltage & Isolation",
    ],
    "EOL Testing": [
        "7.1.1 Check for Abnormal temp & voltages and cell imbalance, Isolation resistance (EOL: 300 amp charge and discharge)",
    ],
    "Pack Assembly": [
        "8.1.1 Check whether PRV nut interfering",
        "8.2.1 Check sealant applied uniformly & MSM gap maintained between two modules (11.2 mm)",
        "8.3.1 Check sealant application and side foam in fire arrestor. Check width 168.4 mm & leak check",
        "8.4.1 Torque check @ M4 Allen head",
        "8.5.1 No gaps allowed / proper fitment check. PRV: M5 - 1 Nm",
        "8.6.1 Terminal Placement verification with Jig",
        "8.6.2 PCB Placement verification with Jig",
        "8.6.3 Stud Placement and weld check with Jig",
        "8.6.4 Mounting point alignment (Heatsink) with Jig",
        "8.6.5 Pack dimensions (430 × 168.4 × 301.35 mm) check",
        "8.6.6 Torque checks & torque marking",
        "8.6.7 Flatness ≤ 0.5 mm",
        "8.6.8 Sealing - PCB, Overall Pack Body & Terminals",
        "8.6.9 Overall aesthetics/cleanliness of the pack",
        "8.6.10 Top casing torque check and paint marking (M4 socket head)",
        "8.6.11 Leak test / Pressure test (Passed/Failed)",
        "8.6.12 Pack weight (CAD Wt.: 38 Kg)",
        "8.7.1 Pack voltage: 100 V Nominal",
        "8.7.2 Isolation resistance: Min in MΩ",
        "8.7.3 Hard leakage to body - Shouldn't present",
        "8.7.4 Voltage taps and thermistor readings with respect to PRAVAIG BMS (PCB communications: voltage + temperature)",
        "8.7.5 Hard leakage to body - Shouldn't present (after bench drop test ~20°)",
    ],
    "Ready for Dispatch": [
        "9.1 Overall pack visual inspection: No defects/no dents/no stains. HV terminals covered with Kapton, PCB covered, PRV placed and covered, product labels, warning labels, +ve & -ve terminal labels, manufacturing date and QC decals placed. (SOC % before dispatch ≤30%)",
    ],
}


def _split_names(combined: str) -> list:
    """Split a name string on ',' returning stripped non-empty parts."""
    parts = [p.strip() for p in combined.split(',') if p.strip()]
    return parts if parts else [combined.strip()]


def format_module_names(combined: str) -> str:
    """Convert 'NameX, NameY' → 'Module X: NameX\nModule Y: NameY'.
    For single name, shows same name for both modules.
    For 3+ names, Module Y shows all names after the first joined by ', '."""
    if not combined or not combined.strip():
        return ''
    parts = _split_names(combined)
    x = parts[0]
    y = ', '.join(parts[1:]) if len(parts) > 1 else x
    return f"Module X: {x}\nModule Y: {y}"


def first_format_module_names(combined: str) -> str:
    """Return only the Module X name (first part, or the single name)."""
    if not combined or not combined.strip():
        return ''
    parts = _split_names(combined)
    return parts[0]

    return combined.strip()


def format_date_str(date_val) -> str:
    """Format a date value as string without microseconds.
    Handles both datetime objects and strings from the database."""
    if not date_val:
        return ""
    if isinstance(date_val, str):
        # Strip microseconds from strings like '2026-01-20 10:28:44.574484'
        dot_pos = date_val.rfind('.')
        if dot_pos != -1 and dot_pos > 10:  # Only strip if it looks like microseconds after a time
            return date_val[:dot_pos]
        return date_val
    # datetime object
    return date_val.strftime("%Y-%m-%d %H:%M:%S")


def safe_write_cell(ws, row: int, column: int, value, font=None):
    """Safely write to a cell, handling merged cells. Optionally override font."""
    try:
        cell = ws.cell(row=row, column=column)
        # Check if this is a MergedCell
        if isinstance(cell, openpyxl.cell.cell.MergedCell):
            # Find the merged range this cell belongs to
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    # Get the top-left cell of the merged range
                    min_row = merged_range.min_row
                    min_col = merged_range.min_col
                    top_left_cell = ws.cell(row=min_row, column=min_col)
                    top_left_cell.value = value
                    if font:
                        top_left_cell.font = font
                    return
        else:
            # Normal cell, write directly
            cell.value = value
            if font:
                cell.font = font
    except Exception as e:
        logger.error(f"Error writing to cell ({row}, {column}): {e}")
        raise


def generate_battery_excel(battery_pack_id: str) -> Optional[Path]:
    """
    Generate individual Excel file for a battery pack
    Uses EXACT same format and row mapping as before

    Returns: Path to generated Excel file
    """
    try:
        # Load template (NEVER modified - always clean)
        if not TEMPLATE_PATH.exists():
            logger.error(f"Template {TEMPLATE_PATH} not found")
            return None

        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        ws = wb.worksheets[0]  # Use first sheet as template

        # Write Battery Pack ID to cell J6 (exactly as before)
        safe_write_cell(ws, 6, 10, battery_pack_id)

        # Write "Pack: XX  Module: SN1 & SN2" to cell P6
        _pi = get_battery_pack_info(battery_pack_id)
        safe_write_cell(ws, 6, 16, f"Pack: {battery_pack_id}  Module: {_pi.get('module_sn1','')} & {_pi.get('module_sn2','')}")

        # Get all QC checks for this battery pack from database
        all_checks = get_qc_checks(battery_pack_id)

        # Group checks by process
        checks_by_process = {}
        for check in all_checks:
            process_name = check['process_name']
            if process_name not in checks_by_process:
                checks_by_process[process_name] = []
            checks_by_process[process_name].append(check)

        # Write data to Excel using EXACT same logic as before
        for process_name, checks in checks_by_process.items():
            # Get process mapping (same as before)
            if process_name not in PROCESS_ROW_MAPPING:
                logger.warning(f"Process '{process_name}' not in mapping")
                continue

            process_info = PROCESS_ROW_MAPPING[process_name]
            start_row = process_info["start_row"]
            process_type = process_info["type"]

            # Write data based on process type (EXACT same logic as app_unified.py)
            if process_type == "standard":
                # Processes 1-7: Cell sorting through EOL Testing
                # Columns: L(12), M(13), N(14), O(15), P(16), Q(17), R(18)
                order = QC_CHECKS_ORDER.get(process_name, [])
                for check in checks:
                    idx = order.index(check['check_name']) if check['check_name'] in order else len(order)
                    row = start_row + idx

                    start_date_str = format_date_str(check.get('start_date'))
                    end_date_str = format_date_str(check.get('end_date'))

                    safe_write_cell(ws, row, 12, check.get('module_x', ''), font=STANDARD_FONT)  # L: Module X QC Result
                    safe_write_cell(ws, row, 13, check.get('module_y', ''), font=STANDARD_FONT)  # M: Module Y QC Result
                    safe_write_cell(ws, row, 14, start_date_str)                 # N: Start date
                    safe_write_cell(ws, row, 15, end_date_str)                   # O: End date
                    safe_write_cell(ws, row, 16, format_module_names(check.get('technician_name', '')))  # P: Technician Name/sign
                    safe_write_cell(ws, row, 17, format_module_names(check.get('qc_name', '')))      # Q: QC Name/Sign
                    safe_write_cell(ws, row, 18, check.get('remarks', ''))      # R: Remarks

            elif process_type == "pack":
                # Process 8: Pack Assembly
                # N and O are merged in template → single Date cell; P and Q merged → single Name cell
                # Write end_date if available, else start_date
                order = QC_CHECKS_ORDER.get(process_name, [])
                for check in checks:
                    idx = order.index(check['check_name']) if check['check_name'] in order else len(order)
                    row = start_row + idx

                    pack_result = check.get('module_x', '')
                    if pack_result == '' or pack_result == 'N/A':
                        pack_result = check.get('module_y', '')

                    date_str = format_date_str(check.get('end_date') or check.get('start_date'))

                    safe_write_cell(ws, row, 12, pack_result, font=STANDARD_FONT)   # L: Pack QC Result
                    safe_write_cell(ws, row, 14, date_str)                          # N-O merged: Date
                    safe_write_cell(ws, row, 16, format_module_names(check.get('technician_name', '')))  # P-Q merged: Name
                    safe_write_cell(ws, row, 18, check.get('remarks', ''))          # R: Remarks

            elif process_type == "dispatch":
                # Process 9-10: Ready for Dispatch
                # N-O merged, P-Q merged → single Date and Name cells
                if len(checks) == 1:
                    row = 62
                    check = checks[0]

                    result = check.get('module_x', '')
                    if result == '' or result == 'N/A':
                        result = check.get('module_y', '')

                    date_str = format_date_str(check.get('end_date') or check.get('start_date'))

                    safe_write_cell(ws, row, 12, result, font=STANDARD_FONT)        # L: Result
                    safe_write_cell(ws, row, 14, date_str)                          # N-O merged: Date
                    safe_write_cell(ws, row, 16, format_module_names(check.get('technician_name', '')))  # P-Q merged: Name
                    safe_write_cell(ws, row, 18, check.get('remarks', ''))          # R: Remarks
                else:
                    # Process 10: Packaging Instructions & PDIR Acceptance
                    # Special: Inspector name in F63, Date in J63
                    if checks:
                        first_check = checks[0]
                        timestamp_str = format_date_str(first_check.get('start_date'))

                        safe_write_cell(ws, 63, 6, first_format_module_names(first_check.get('qc_name', '')))  # F63: Inspector Name
                        safe_write_cell(ws, 63, 10, timestamp_str)                  # J63: Date

                    # Data rows starting at 64
                    for idx, check in enumerate(checks):
                        row = 64 + idx

                        result = check.get('module_x', '')
                        if result == '' or result == 'N/A':
                            result = check.get('module_y', '')

                        safe_write_cell(ws, row, 12, result, font=STANDARD_FONT)    # L: Result
                        safe_write_cell(ws, row, 16, check.get('remarks', ''))      # P: Comments

        # Save to individual file
        output_dir = Path("excel_reports")
        output_dir.mkdir(exist_ok=True)

        output_path = output_dir / f"{battery_pack_id}.xlsx"
        wb.save(output_path)

        logger.info(f"Generated Excel: {output_path}")
        return output_path

    except Exception as e:
        logger.error(f"Error generating Excel for {battery_pack_id}: {e}", exc_info=True)
        return None


def generate_master_excel() -> Optional[Path]:
    """
    Generate master sample.xlsx with all battery packs as sheets
    EXACT same format as before - just generated from database

    Returns: Path to master Excel file
    """
    try:
        if not TEMPLATE_PATH.exists():
            logger.error(f"Template {TEMPLATE_PATH} not found")
            return None

        # Load clean template (NEVER modified)
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

        # Get all battery pack IDs from database
        battery_ids = get_all_battery_packs()

        if not battery_ids:
            logger.info("No battery packs in database yet")
            return MASTER_OUTPUT_PATH

        # Remove all sheets except template (first sheet)
        while len(wb.worksheets) > 1:
            wb.remove(wb.worksheets[1])

        # Create sheet for each battery pack
        for battery_id in battery_ids:
            # Copy template sheet
            template_sheet = wb.worksheets[0]
            ws = wb.copy_worksheet(template_sheet)
            ws.title = battery_id

            # Write Battery Pack ID
            safe_write_cell(ws, 6, 10, battery_id)

            # Write "Pack: XX  Module: SN1 & SN2" to cell P6
            _pi = get_battery_pack_info(battery_id)
            safe_write_cell(ws, 6, 16, f"Pack: {battery_id}  Module: {_pi.get('module_sn1','')} & {_pi.get('module_sn2','')}")

            # Get all QC checks for this battery pack
            all_checks = get_qc_checks(battery_id)

            # Group by process
            checks_by_process = {}
            for check in all_checks:
                process_name = check['process_name']
                if process_name not in checks_by_process:
                    checks_by_process[process_name] = []
                checks_by_process[process_name].append(check)

            # Write data (EXACT same logic as individual file)
            for process_name, checks in checks_by_process.items():
                if process_name not in PROCESS_ROW_MAPPING:
                    continue

                process_info = PROCESS_ROW_MAPPING[process_name]
                start_row = process_info["start_row"]
                process_type = process_info["type"]

                # Same writing logic as generate_battery_excel (keeping code identical)
                if process_type == "standard":
                    order = QC_CHECKS_ORDER.get(process_name, [])
                    for check in checks:
                        idx = order.index(check['check_name']) if check['check_name'] in order else len(order)
                        row = start_row + idx

                        start_date_str = format_date_str(check.get('start_date'))
                        end_date_str = format_date_str(check.get('end_date'))

                        safe_write_cell(ws, row, 12, check.get('module_x', ''), font=STANDARD_FONT)
                        safe_write_cell(ws, row, 13, check.get('module_y', ''), font=STANDARD_FONT)
                        safe_write_cell(ws, row, 14, start_date_str)
                        safe_write_cell(ws, row, 15, end_date_str)
                        safe_write_cell(ws, row, 16, format_module_names(check.get('technician_name', '')))
                        safe_write_cell(ws, row, 17, format_module_names(check.get('qc_name', '')))
                        safe_write_cell(ws, row, 18, check.get('remarks', ''))

                elif process_type == "pack":
                    order = QC_CHECKS_ORDER.get(process_name, [])
                    for check in checks:
                        idx = order.index(check['check_name']) if check['check_name'] in order else len(order)
                        row = start_row + idx
                        pack_result = check.get('module_x', '')
                        if pack_result == '' or pack_result == 'N/A':
                            pack_result = check.get('module_y', '')

                        safe_write_cell(ws, row, 12, pack_result, font=STANDARD_FONT)
                        safe_write_cell(ws, row, 14, format_date_str(check.get('end_date') or check.get('start_date')))
                        safe_write_cell(ws, row, 16, format_module_names(check.get('technician_name', '')))
                        safe_write_cell(ws, row, 18, check.get('remarks', ''))

                elif process_type == "dispatch":
                    if len(checks) == 1:
                        row = 62
                        check = checks[0]
                        result = check.get('module_x', '')
                        if result == '' or result == 'N/A':
                            result = check.get('module_y', '')

                        safe_write_cell(ws, row, 12, result, font=STANDARD_FONT)
                        safe_write_cell(ws, row, 14, format_date_str(check.get('end_date') or check.get('start_date')))
                        safe_write_cell(ws, row, 16, format_module_names(check.get('technician_name', '')))
                        safe_write_cell(ws, row, 18, check.get('remarks', ''))
                    else:
                        if checks:
                            first_check = checks[0]
                            timestamp_str = format_date_str(first_check.get('start_date'))

                            safe_write_cell(ws, 63, 6, first_format_module_names(first_check.get('qc_name', '')))
                            safe_write_cell(ws, 63, 10, timestamp_str)

                        for idx, check in enumerate(checks):
                            row = 64 + idx
                            result = check.get('module_x', '')
                            if result == '' or result == 'N/A':
                                result = check.get('module_y', '')
                            safe_write_cell(ws, row, 12, result, font=STANDARD_FONT)
                            safe_write_cell(ws, row, 16, check.get('remarks', ''))

        # Save master output file (NOT the template!)
        wb.save(MASTER_OUTPUT_PATH)

        logger.info(f"Generated master Excel: {MASTER_OUTPUT_PATH} with {len(battery_ids)} battery packs")
        return MASTER_OUTPUT_PATH

    except Exception as e:
        logger.error(f"Error generating master Excel: {e}", exc_info=True)
        return None


def update_excel_after_entry(battery_pack_id: str):
    """
    Update Excel file for the individual battery pack immediately after data entry.
    Master Excel (sample.xlsx) is intentionally NOT regenerated here — it is
    generated on-demand when the Reports tab is used. This keeps saves fast.
    """
    try:
        # Generate individual Excel file only
        generate_battery_excel(battery_pack_id)
        logger.info(f"Excel updated for {battery_pack_id}")

    except Exception as e:
        logger.error(f"Error updating Excel: {e}", exc_info=True)


# ============================================================================
# ON-DEMAND EXCEL GENERATION (Returns bytes - no file saving)
# These functions are for Reports tab downloads only
# They do NOT affect the production process
# ============================================================================

def generate_battery_excel_bytes(battery_pack_id: str) -> Optional[bytes]:
    """
    Generate Excel for a single battery pack and return as bytes (in-memory).
    Does NOT save to disk - for download only.
    """
    try:
        # Load clean template (NEVER modified)
        if not TEMPLATE_PATH.exists():
            logger.error(f"Template {TEMPLATE_PATH} not found")
            return None

        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        ws = wb.worksheets[0]

        # Write Battery Pack ID
        safe_write_cell(ws, 6, 10, battery_pack_id)

        # Write "Pack: XX  Module: SN1 & SN2" to cell P6
        _pi = get_battery_pack_info(battery_pack_id)
        safe_write_cell(ws, 6, 16, f"Pack: {battery_pack_id}  Module: {_pi.get('module_sn1','')} & {_pi.get('module_sn2','')}")

        # Get all QC checks for this battery pack
        all_checks = get_qc_checks(battery_pack_id)

        # Group checks by process
        checks_by_process = {}
        for check in all_checks:
            process_name = check['process_name']
            if process_name not in checks_by_process:
                checks_by_process[process_name] = []
            checks_by_process[process_name].append(check)

        # Write data (same logic as generate_battery_excel)
        for process_name, checks in checks_by_process.items():
            if process_name not in PROCESS_ROW_MAPPING:
                continue

            process_info = PROCESS_ROW_MAPPING[process_name]
            start_row = process_info["start_row"]
            process_type = process_info["type"]

            if process_type == "standard":
                order = QC_CHECKS_ORDER.get(process_name, [])
                for check in checks:
                    idx = order.index(check['check_name']) if check['check_name'] in order else len(order)
                    row = start_row + idx

                    start_date_str = format_date_str(check.get('start_date'))
                    end_date_str = format_date_str(check.get('end_date'))

                    safe_write_cell(ws, row, 12, check.get('module_x', ''), font=STANDARD_FONT)
                    safe_write_cell(ws, row, 13, check.get('module_y', ''), font=STANDARD_FONT)
                    safe_write_cell(ws, row, 14, start_date_str)
                    safe_write_cell(ws, row, 15, end_date_str)
                    safe_write_cell(ws, row, 16, format_module_names(check.get('technician_name', '')))
                    safe_write_cell(ws, row, 17, format_module_names(check.get('qc_name', '')))
                    safe_write_cell(ws, row, 18, check.get('remarks', ''))

            elif process_type == "pack":
                order = QC_CHECKS_ORDER.get(process_name, [])
                for check in checks:
                    idx = order.index(check['check_name']) if check['check_name'] in order else len(order)
                    row = start_row + idx
                    pack_result = check.get('module_x', '')
                    if pack_result == '' or pack_result == 'N/A':
                        pack_result = check.get('module_y', '')

                    safe_write_cell(ws, row, 12, pack_result, font=STANDARD_FONT)
                    safe_write_cell(ws, row, 14, format_date_str(check.get('end_date') or check.get('start_date')))
                    safe_write_cell(ws, row, 16, format_module_names(check.get('technician_name', '')))
                    safe_write_cell(ws, row, 18, check.get('remarks', ''))

            elif process_type == "dispatch":
                if len(checks) == 1:
                    row = 62
                    check = checks[0]
                    result = check.get('module_x', '')
                    if result == '' or result == 'N/A':
                        result = check.get('module_y', '')

                    safe_write_cell(ws, row, 12, result, font=STANDARD_FONT)
                    safe_write_cell(ws, row, 14, format_date_str(check.get('end_date') or check.get('start_date')))
                    safe_write_cell(ws, row, 16, format_module_names(check.get('technician_name', '')))
                    safe_write_cell(ws, row, 18, check.get('remarks', ''))
                else:
                    if checks:
                        first_check = checks[0]
                        timestamp_str = format_date_str(first_check.get('start_date'))
                        safe_write_cell(ws, 63, 6, first_format_module_names(first_check.get('qc_name', '')))
                        safe_write_cell(ws, 63, 10, timestamp_str)

                    for idx, check in enumerate(checks):
                        row = 64 + idx
                        result = check.get('module_x', '')
                        if result == '' or result == 'N/A':
                            result = check.get('module_y', '')
                        safe_write_cell(ws, row, 12, result, font=STANDARD_FONT)
                        safe_write_cell(ws, row, 16, check.get('remarks', ''))

        # Save to bytes (in-memory)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    except Exception as e:
        logger.error(f"Error generating Excel bytes for {battery_pack_id}: {e}", exc_info=True)
        return None


def generate_all_reports_excel_bytes() -> Optional[bytes]:
    """
    Generate Excel with all battery packs as sheets and return as bytes (in-memory).
    Does NOT save to disk - for download only.
    """
    try:
        if not TEMPLATE_PATH.exists():
            logger.error(f"Template {TEMPLATE_PATH} not found")
            return None

        wb = openpyxl.load_workbook(TEMPLATE_PATH)

        # Get all battery pack IDs from database
        battery_ids = get_all_battery_packs()

        if not battery_ids:
            logger.info("No battery packs in database")
            return None

        # Remove all sheets except template (first sheet)
        while len(wb.worksheets) > 1:
            wb.remove(wb.worksheets[1])

        # Create sheet for each battery pack
        for battery_id in battery_ids:
            template_sheet = wb.worksheets[0]
            ws = wb.copy_worksheet(template_sheet)
            ws.title = battery_id

            # Write Battery Pack ID
            safe_write_cell(ws, 6, 10, battery_id)

            # Write "Pack: XX  Module: SN1 & SN2" to cell P6
            _pi = get_battery_pack_info(battery_id)
            safe_write_cell(ws, 6, 16, f"Pack: {battery_id}  Module: {_pi.get('module_sn1','')} & {_pi.get('module_sn2','')}")

            # Get all QC checks for this battery pack
            all_checks = get_qc_checks(battery_id)

            # Group by process
            checks_by_process = {}
            for check in all_checks:
                process_name = check['process_name']
                if process_name not in checks_by_process:
                    checks_by_process[process_name] = []
                checks_by_process[process_name].append(check)

            # Write data (same logic as generate_master_excel)
            for process_name, checks in checks_by_process.items():
                if process_name not in PROCESS_ROW_MAPPING:
                    continue

                process_info = PROCESS_ROW_MAPPING[process_name]
                start_row = process_info["start_row"]
                process_type = process_info["type"]

                if process_type == "standard":
                    order = QC_CHECKS_ORDER.get(process_name, [])
                    for check in checks:
                        idx = order.index(check['check_name']) if check['check_name'] in order else len(order)
                        row = start_row + idx

                        start_date_str = format_date_str(check.get('start_date'))
                        end_date_str = format_date_str(check.get('end_date'))

                        safe_write_cell(ws, row, 12, check.get('module_x', ''), font=STANDARD_FONT)
                        safe_write_cell(ws, row, 13, check.get('module_y', ''), font=STANDARD_FONT)
                        safe_write_cell(ws, row, 14, start_date_str)
                        safe_write_cell(ws, row, 15, end_date_str)
                        safe_write_cell(ws, row, 16, format_module_names(check.get('technician_name', '')))
                        safe_write_cell(ws, row, 17, format_module_names(check.get('qc_name', '')))
                        safe_write_cell(ws, row, 18, check.get('remarks', ''))

                elif process_type == "pack":
                    order = QC_CHECKS_ORDER.get(process_name, [])
                    for check in checks:
                        idx = order.index(check['check_name']) if check['check_name'] in order else len(order)
                        row = start_row + idx
                        pack_result = check.get('module_x', '')
                        if pack_result == '' or pack_result == 'N/A':
                            pack_result = check.get('module_y', '')

                        safe_write_cell(ws, row, 12, pack_result, font=STANDARD_FONT)
                        safe_write_cell(ws, row, 14, format_date_str(check.get('end_date') or check.get('start_date')))
                        safe_write_cell(ws, row, 16, format_module_names(check.get('technician_name', '')))
                        safe_write_cell(ws, row, 18, check.get('remarks', ''))

                elif process_type == "dispatch":
                    if len(checks) == 1:
                        row = 62
                        check = checks[0]
                        result = check.get('module_x', '')
                        if result == '' or result == 'N/A':
                            result = check.get('module_y', '')

                        safe_write_cell(ws, row, 12, result, font=STANDARD_FONT)
                        safe_write_cell(ws, row, 14, format_date_str(check.get('end_date') or check.get('start_date')))
                        safe_write_cell(ws, row, 16, format_module_names(check.get('technician_name', '')))
                        safe_write_cell(ws, row, 18, check.get('remarks', ''))
                    else:
                        if checks:
                            first_check = checks[0]
                            timestamp_str = format_date_str(first_check.get('start_date'))
                            safe_write_cell(ws, 63, 6, first_format_module_names(first_check.get('qc_name', '')))
                            safe_write_cell(ws, 63, 10, timestamp_str)

                        for idx, check in enumerate(checks):
                            row = 64 + idx
                            result = check.get('module_x', '')
                            if result == '' or result == 'N/A':
                                result = check.get('module_y', '')
                            safe_write_cell(ws, row, 12, result, font=STANDARD_FONT)
                            safe_write_cell(ws, row, 16, check.get('remarks', ''))

        # Save to bytes (in-memory)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(f"Generated all reports Excel bytes with {len(battery_ids)} battery packs")
        return output.getvalue()

    except Exception as e:
        logger.error(f"Error generating all reports Excel bytes: {e}", exc_info=True)
        return None

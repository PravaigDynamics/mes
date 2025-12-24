import openpyxl
from pathlib import Path

# Load sample.xlsx
wb = openpyxl.load_workbook("sample.xlsx")
ws = wb.worksheets[0]

print(f"Sheet name: {ws.title}")
print(f"Max row: {ws.max_row}")
print(f"Max column: {ws.max_column}")
print("\nMerged cells:")
for merged in ws.merged_cells.ranges:
    print(f"  {merged}")

print("\n\nFirst 50 rows structure:")
print("-" * 150)

for row_idx in range(1, min(51, ws.max_row + 1)):
    row_data = []
    for col_idx in range(1, min(15, ws.max_column + 1)):
        cell = ws.cell(row=row_idx, column=col_idx)
        if isinstance(cell, openpyxl.cell.cell.MergedCell):
            row_data.append("[MERGED]")
        else:
            value = str(cell.value) if cell.value is not None else ""
            if len(value) > 20:
                value = value[:17] + "..."
            row_data.append(value)

    # Only print rows that have some content
    if any(val and val != "[MERGED]" for val in row_data):
        print(f"Row {row_idx:3d}: {' | '.join(row_data)}")

print("\n\nColumn headers (looking for Module X, Module Y, etc.):")
print("-" * 150)

# Find header row
for row_idx in range(1, 30):
    for col_idx in range(1, 15):
        cell = ws.cell(row=row_idx, column=col_idx)
        if not isinstance(cell, openpyxl.cell.cell.MergedCell):
            val = str(cell.value) if cell.value else ""
            if "Module X" in val or "Module Y" in val or "Technician" in val:
                print(f"Row {row_idx}, Col {col_idx}: {val}")

wb.close()

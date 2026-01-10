"""
Battery Pack Manufacturing Execution System
Professional Enterprise Application

MODIFIED VERSION - Uses Database + Excel Export
- Same Excel format and data mapping
- Same UI and functionality
- Better reliability for concurrent users
- Excel files generated from database
"""

import sys
import io
import os
import time
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional
import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import openpyxl
import qrcode
from PIL import Image
import base64

# Import database and Excel generator
from database import (
    init_database, save_qc_checks, update_process_completion,
    check_process_status, battery_pack_exists, get_all_battery_packs,
    get_qc_checks, get_dashboard_status
)
from excel_generator import (
    generate_battery_excel, generate_master_excel, update_excel_after_entry
)
from backup_manager import create_backup, list_backups, get_database_size

# Setup basic logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Initialize database on startup
try:
    init_database()
    logger.info("Database initialized")

except Exception as e:
    logger.error(f"Database initialization failed: {e}")

# Process definitions - Updated with actual QC checks from template
PROCESS_DEFINITIONS = {
    "Cell sorting": {
        "qc_checks": [
            "Acceptable range Voltage: 3.58 to 3.6 V, IR: 11.5 to 12.5 mΩ",
            "Jig setup on the flat surface",
            "Visual & spacing inspection"
        ]
    },
    "Module assembly": {
        "qc_checks": [
            "Aluminum base plate fixing in Base plate fixture & stitch nut",
            "Adhesive dispensing on cells placed in the jig & lio of cell assembly jig on baseplate fixture after curing",
            "Dispensing the cell assembly jig & fixture",
            "Fixing of Downwell (PC sheets) in the cell assembly",
            "Thermistor & voltage tagging on Cell assembly(Routing & Placement)",
            "Busbar fitment on CPT (adhesive dispensing & voltage tap,thermistor sub assembly)"
        ]
    },
    "Pre Encapsulation": {
        "qc_checks": [
            "The 90 % of encapsulation on module by tightening the module independently using the holding fixture",
            "Bus bar top assembly placement and Soldering of voltage taps on the cells",
            "Stripping crimping & connector insertion for voltage tap and thermistor"
        ]
    },
    "Wire Bonding": {
        "qc_checks": [
            "As per the cell assembly procedure"
        ]
    },
    "Post Encapsulation": {
        "qc_checks": [
            "The remaining 20 % of encapsulation on wire bonded module",
            "Dimensions of the part"
        ]
    },
    "EOL Testing": {
        "qc_checks": [
            "Check for Abnormal temp & voltages and cell imbalance, Isolation resistance (EOL: 300 amp charge and discharge)"
        ]
    },
    "Pack assembly": {
        "qc_checks": [
            "Pack Assembly(Mock fitment of module 1 & 2 in base plate enclosure)",
            "Fitment of module 1 & module 2between MSM sandwich bodies on the base plate with sealant and screw with M6 allen head",
            "Assembly of base plates (left & flame arrester sizes (M3) with sealant /foam and M6)",
            "Assembly of busbar series with UX Allen head",
            "Cell box top cover assembly- PCB joining ;PRV assembly; busbar/array & cleanliness",
            "Final QC on the pack level with CTQs",
            "Sealing- PCB, Overall Pack Body & Terminals",
            "Overall aesthetics/cleanliness of the pack",
            "Pre-casing torque check and paint marking (M4 socket head)",
            "Leak test (Pressure test) (chaser/Fixer)",
            "Labelling of the battery pack",
            "Torque checks & torque marking",
            "Sealing - PCB, Overall Pack Body & Terminals)",
            "Overall aesthetics/cleanliness of the pack",
            "Pre-casing torque check and paint marking (M4 socket head)",
            "Leak test (Pressure test) (chaser/Fixer)",
            "Labelling of the battery pack",
            "Torque checks & torque marking",
            "Hard leakage to body - Shouldn't present",
            "Voltage and thermaltor readings with respect to PREVAID BMS (PCB communications - voltage + temperature)",
            "Hard leakage to body: Shouldn't present",
            "Isolation resistance: Min in Mohm"
        ]
    },
    "Ready for Dispatch": {
        "qc_checks": [
            "Overall pack visual inspection: No defects/no dents/no stains, HV terminals covered, PCB covered, PRV placed, labels verified"
        ]
    }
}

# Row mapping for Excel sheet
# Maps process name to starting row and column configuration
PROCESS_ROW_MAPPING = {
    "Cell sorting": {"start_row": 8, "type": "standard"},  # Processes 1-7
    "Module assembly": {"start_row": 11, "type": "standard"},
    "Pre Encapsulation": {"start_row": 14, "type": "standard"},
    "Wire Bonding": {"start_row": 17, "type": "standard"},
    "Post Encapsulation": {"start_row": 20, "type": "standard"},
    "EOL Testing": {"start_row": 38, "type": "standard"},
    "Pack assembly": {"start_row": 40, "type": "pack"},  # Process 8
    "Ready for Dispatch": {"start_row": 62, "type": "dispatch"}  # Processes 9-10
}

# Helper Functions
def safe_write_cell(ws, row: int, column: int, value):
    """Safely write to a cell, handling merged cells."""
    try:
        cell = ws.cell(row=row, column=column)
        # Check if this is a MergedCell
        if isinstance(cell, openpyxl.cell.cell.MergedCell):
            # Find the merged range this cell belongs to
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    # Get the top-left cell of the merged range
                    min_row = merged_range.min_row
                    min_col = merged_range.min_col
                    top_left_cell = ws.cell(row=min_row, column=min_col)
                    top_left_cell.value = value
                    return
        else:
            # Normal cell, write directly
            cell.value = value
    except Exception as e:
        logger.error(f"Error writing to cell ({row}, {column}): {e}")
        raise

def check_battery_exists(battery_pack_id: str) -> dict:
    """Check if battery pack ID already exists in system."""
    exists_info = {
        'qr_exists': False,
        'data_exists': False,
        'qr_path': None,
        'sheet_exists': False
    }

    # Check if QR code file exists
    qr_dir = Path("qr_codes")
    qr_dir.mkdir(exist_ok=True)
    qr_path = qr_dir / f"{battery_pack_id}.png"

    if qr_path.exists():
        exists_info['qr_exists'] = True
        exists_info['qr_path'] = qr_path

    # Check if battery exists in database
    exists_info['data_exists'] = battery_pack_exists(battery_pack_id)
    exists_info['sheet_exists'] = exists_info['data_exists']

    return exists_info

def check_process_data_exists(battery_pack_id: str, process_name: str) -> dict:
    """
    Check if data already exists for a specific process in a battery pack.
    NOW READS FROM DATABASE instead of Excel
    """
    return check_process_status(battery_pack_id, process_name)

def complete_process(battery_pack_id: str, process_name: str) -> Path:
    """
    Complete a process by updating the end date.
    NOW: Update database + regenerate Excel
    """
    try:
        # Update database
        success = update_process_completion(battery_pack_id, process_name)

        if not success:
            raise ValueError(f"Failed to update database for {battery_pack_id}")

        # Regenerate Excel files immediately
        update_excel_after_entry(battery_pack_id)

        # Automatic backup after process completion
        try:
            backup_file = create_backup()
            if backup_file:
                logger.info(f"Automatic backup created after completing {process_name} for {battery_pack_id}")
        except Exception as backup_error:
            logger.warning(f"Auto-backup after process completion failed: {backup_error}")

        logger.info(f"Completed process {process_name} for {battery_pack_id}")
        return Path("sample.xlsx")

    except Exception as e:
        logger.error(f"Error completing process: {e}", exc_info=True)
        raise

def generate_qr_code(battery_pack_id: str, size: int = 300, include_label: bool = True) -> bytes:
    """Generate QR code for battery pack ID and save to qr_codes folder."""
    try:
        # Create QR code
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_H,
            box_size=10,
            border=4,
        )

        # Generate URL from environment variable (HTTPS with domain for camera support)
        base_url = os.getenv('APP_BASE_URL', 'https://mes.pravaig.com')
        data = f"{base_url}/entry/{battery_pack_id}"
        qr.add_data(data)
        qr.make(fit=True)

        # Create image
        img = qr.make_image(fill_color="black", back_color="white")

        # Resize to requested size
        img = img.resize((size, size), Image.Resampling.LANCZOS)

        # Add label if requested
        if include_label:
            from PIL import ImageDraw, ImageFont
            # Create new image with space for label
            label_height = 40
            new_img = Image.new('RGB', (size, size + label_height), 'white')
            new_img.paste(img, (0, 0))

            # Draw label
            draw = ImageDraw.Draw(new_img)
            text = battery_pack_id
            # Use default font
            try:
                font = ImageFont.truetype("arial.ttf", 20)
            except:
                font = ImageFont.load_default()

            # Get text bbox and center it
            bbox = draw.textbbox((0, 0), text, font=font)
            text_width = bbox[2] - bbox[0]
            text_x = (size - text_width) // 2
            text_y = size + 10

            draw.text((text_x, text_y), text, fill='black', font=font)
            img = new_img

        # Save to qr_codes folder
        qr_dir = Path("qr_codes")
        qr_dir.mkdir(exist_ok=True)
        qr_path = qr_dir / f"{battery_pack_id}.png"
        img.save(qr_path)
        logger.info(f"QR code saved to {qr_path}")

        # Convert to bytes for display
        img_buffer = io.BytesIO()
        img.save(img_buffer, format='PNG')
        img_buffer.seek(0)
        return img_buffer.getvalue()

    except Exception as e:
        logger.error(f"QR generation error: {e}")
        raise

def get_battery_report_path(battery_pack_id: str) -> Optional[Path]:
    """Get the Excel report path for a battery pack."""
    excel_dir = Path("excel_reports")
    if not excel_dir.exists():
        excel_dir.mkdir(parents=True, exist_ok=True)

    # Look for existing file
    pattern = f"{battery_pack_id}*.xlsx"
    existing_files = list(excel_dir.glob(pattern))

    if existing_files:
        return existing_files[0]

    return excel_dir / f"{battery_pack_id}.xlsx"

def add_detailed_entry(battery_pack_id: str, process_name: str, technician_name: str,
                      qc_name: str, remarks: str, checks: List[Dict]) -> Path:
    """
    Add detailed entry - NOW saves to database then generates Excel
    Excel format remains EXACTLY the same
    """
    try:
        # 1. Save to database (handles concurrent writes safely)
        success = save_qc_checks(
            pack_id=battery_pack_id,
            process_name=process_name,
            technician_name=technician_name,
            qc_name=qc_name,
            remarks=remarks,
            checks=checks
        )

        if not success:
            raise ValueError("Failed to save data to database")

        # 2. Generate Excel files immediately (same format as before)
        update_excel_after_entry(battery_pack_id)

        # 3. Automatic backup after data entry
        try:
            backup_file = create_backup()
            if backup_file:
                logger.info(f"Automatic backup created after data entry for {battery_pack_id}")
        except Exception as backup_error:
            logger.warning(f"Auto-backup after data entry failed: {backup_error}")

        logger.info(f"Saved data for {battery_pack_id} - Process: {process_name}")
        return Path("sample.xlsx")

    except Exception as e:
        logger.error(f"Data entry error: {e}", exc_info=True)
        raise

# Continue with EXACT same UI code from original app_unified.py...
# (All the Streamlit UI code below remains UNCHANGED)
# Page configuration
st.set_page_config(
    page_title="Battery Pack MES",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Professional CSS - Modern Enterprise Design
st.markdown("""
<style>
    /* Root Variables - Professional Color Palette */
    :root {
        --primary-color: #1565C0;
        --primary-dark: #0D47A1;
        --primary-light: #1976D2;
        --success-color: #2E7D32;
        --success-light: #66BB6A;
        --error-color: #C62828;
        --error-light: #EF5350;
        --warning-color: #F57C00;
        --info-color: #0288D1;
        --gray-50: #FAFAFA;
        --gray-100: #F5F5F5;
        --gray-200: #EEEEEE;
        --gray-300: #E0E0E0;
        --gray-400: #BDBDBD;
        --gray-500: #9E9E9E;
        --gray-700: #616161;
        --gray-800: #424242;
        --gray-900: #212121;
    }

    /* Global Styles */
    .main > div {
        padding: 1.5rem 2rem;
        max-width: 1400px;
        margin: 0 auto;
    }

    /* Typography */
    h1 {
        color: var(--gray-900);
        font-size: 2rem;
        font-weight: 600;
        margin-bottom: 0.5rem;
        letter-spacing: -0.02em;
    }

    h2 {
        color: var(--gray-800);
        font-size: 1.5rem;
        font-weight: 600;
        margin-top: 2rem;
        margin-bottom: 1rem;
    }

    h3 {
        color: var(--gray-800);
        font-size: 1.25rem;
        font-weight: 500;
        margin-top: 1.5rem;
        margin-bottom: 0.75rem;
    }

    /* Buttons - Material Design Style */
    .stButton > button {
        width: 100%;
        padding: 0.75rem 1.5rem;
        font-size: 0.95rem;
        font-weight: 500;
        border-radius: 4px;
        border: none;
        transition: all 0.2s ease;
        letter-spacing: 0.02em;
        text-transform: uppercase;
    }

    .stButton > button:hover {
        transform: translateY(-1px);
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.15);
    }

    /* Input Fields */
    .stTextInput > div > div > input,
    .stSelectbox > div > div > select,
    .stTextArea > div > div > textarea {
        font-size: 1rem;
        padding: 0.75rem;
        border: 1px solid var(--gray-300);
        border-radius: 4px;
        transition: border-color 0.2s ease;
    }

    .stTextInput > div > div > input:focus,
    .stSelectbox > div > div > select:focus,
    .stTextArea > div > div > textarea:focus {
        border-color: var(--primary-color);
        box-shadow: 0 0 0 2px rgba(21, 101, 192, 0.1);
    }

    /* Labels */
    .stTextInput > label,
    .stSelectbox > label,
    .stTextArea > label {
        font-size: 0.875rem;
        font-weight: 500;
        color: var(--gray-700);
        margin-bottom: 0.5rem;
    }

    /* Tabs - Professional Navigation */
    .stTabs [data-baseweb="tab-list"] {
        gap: 0;
        background-color: var(--gray-100);
        border-bottom: 2px solid var(--gray-300);
        padding: 0;
    }

    .stTabs [data-baseweb="tab"] {
        height: 56px;
        padding: 0 24px;
        background-color: transparent;
        border-radius: 0;
        font-size: 0.875rem;
        font-weight: 500;
        color: var(--gray-700);
        text-transform: uppercase;
        letter-spacing: 0.05em;
        border-bottom: 3px solid transparent;
    }

    .stTabs [aria-selected="true"] {
        background-color: white;
        color: var(--primary-color);
        border-bottom: 3px solid var(--primary-color);
    }

    /* Cards */
    .card {
        background: white;
        border: 1px solid var(--gray-200);
        border-radius: 8px;
        padding: 1.5rem;
        margin: 1rem 0;
        box-shadow: 0 1px 3px rgba(0, 0, 0, 0.06);
    }

    .card-header {
        font-size: 1.125rem;
        font-weight: 600;
        color: var(--gray-800);
        margin-bottom: 1rem;
        padding-bottom: 0.75rem;
        border-bottom: 1px solid var(--gray-200);
    }

    /* Status Badges */
    .badge {
        display: inline-block;
        padding: 0.35rem 0.75rem;
        border-radius: 12px;
        font-size: 0.75rem;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.05em;
    }

    .badge-success {
        background-color: #E8F5E9;
        color: var(--success-color);
    }

    .badge-info {
        background-color: #E3F2FD;
        color: var(--info-color);
    }

    .badge-warning {
        background-color: #FFF3E0;
        color: var(--warning-color);
    }

    .badge-error {
        background-color: #FFEBEE;
        color: var(--error-color);
    }

    /* Alert Boxes */
    .alert {
        padding: 1rem 1.25rem;
        border-radius: 4px;
        margin: 1rem 0;
        border-left: 4px solid;
    }

    .alert-success {
        background-color: #E8F5E9;
        border-left-color: var(--success-color);
        color: var(--success-color);
    }

    .alert-info {
        background-color: #E3F2FD;
        border-left-color: var(--info-color);
        color: #01579B;
    }

    .alert-warning {
        background-color: #FFF3E0;
        border-left-color: var(--warning-color);
        color: #E65100;
    }

    .alert-error {
        background-color: #FFEBEE;
        border-left-color: var(--error-color);
        color: var(--error-color);
    }

    /* Scanner Method Cards */
    .method-card {
        background: white;
        border: 2px solid var(--gray-200);
        border-radius: 8px;
        padding: 1.5rem;
        text-align: center;
        cursor: pointer;
        transition: all 0.2s ease;
        height: 100%;
    }

    .method-card:hover {
        border-color: var(--primary-color);
        box-shadow: 0 4px 12px rgba(21, 101, 192, 0.15);
    }

    .method-card.primary {
        border-color: var(--primary-color);
        background: linear-gradient(135deg, #E3F2FD 0%, #BBDEFB 100%);
    }

    .method-card.secondary {
        border-color: var(--success-color);
        background: linear-gradient(135deg, #E8F5E9 0%, #C8E6C9 100%);
    }

    .method-title {
        font-size: 1.125rem;
        font-weight: 600;
        color: var(--gray-900);
        margin-bottom: 0.5rem;
    }

    .method-description {
        font-size: 0.875rem;
        color: var(--gray-600);
        line-height: 1.5;
    }

    /* Metrics Cards */
    .metric-card {
        background: white;
        border: 1px solid var(--gray-200);
        border-radius: 8px;
        padding: 1.5rem;
        text-align: center;
    }

    .metric-value {
        font-size: 2rem;
        font-weight: 700;
        color: var(--primary-color);
        line-height: 1;
        margin: 0.5rem 0;
    }

    .metric-label {
        font-size: 0.875rem;
        font-weight: 500;
        color: var(--gray-600);
        text-transform: uppercase;
        letter-spacing: 0.05em;
    }

    /* File Upload */
    .uploadedFile {
        font-size: 0.875rem;
    }

    [data-testid="stFileUploader"] {
        border: 2px dashed var(--gray-300);
        border-radius: 8px;
        padding: 2rem;
        background: var(--gray-50);
        transition: border-color 0.2s ease;
    }

    [data-testid="stFileUploader"]:hover {
        border-color: var(--primary-color);
        background: white;
    }

    /* QR Scanner Container */
    .scanner-container {
        background: white;
        border: 1px solid var(--gray-200);
        border-radius: 8px;
        padding: 1.5rem;
        margin: 1rem 0;
    }

    /* Dividers */
    hr {
        border: none;
        border-top: 1px solid var(--gray-200);
        margin: 2rem 0;
    }

    /* Mobile Responsive */
    @media (max-width: 768px) {
        .main > div {
            padding: 1rem;
        }

        h1 {
            font-size: 1.5rem;
        }

        .method-card {
            padding: 1rem;
        }

        .metric-value {
            font-size: 1.5rem;
        }
    }

    /* Hide Streamlit Branding */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
</style>
""", unsafe_allow_html=True)


# ============================================================================
# QR CODE DETECTION
# ============================================================================

def decode_qr_from_image(image):
    """Decode QR code from uploaded image using OpenCV."""
    try:
        import cv2
        import numpy as np

        # Convert PIL Image to numpy array
        img_array = np.array(image)

        # Convert RGB to BGR (OpenCV uses BGR)
        if len(img_array.shape) == 3 and img_array.shape[2] == 3:
            img_array = cv2.cvtColor(img_array, cv2.COLOR_RGB2BGR)
        elif len(img_array.shape) == 3 and img_array.shape[2] == 4:
            img_array = cv2.cvtColor(img_array, cv2.COLOR_RGBA2BGR)

        # Initialize QR code detector
        detector = cv2.QRCodeDetector()

        # Detect and decode QR code
        data, bbox, straight_qrcode = detector.detectAndDecode(img_array)

        if data:
            logger.info(f"QR code detected: {data}")
            return data
        else:
            logger.warning("No QR code found in image")
            return None

    except Exception as e:
        logger.error(f"QR decode error: {e}", exc_info=True)
        return None


def extract_battery_id_from_url(url):
    """Extract battery pack ID from QR code URL."""
    try:
        if '/entry/' in url:
            battery_id = url.split('/entry/')[-1]
            battery_id = battery_id.split('?')[0]
            battery_id = battery_id.split('#')[0]
            return battery_id.strip()
        return url
    except:
        return url


def check_camera_support():
    """Check if camera scanning is supported in current environment."""
    # Camera works on localhost and HTTPS
    # For HTTP deployments, recommend photo upload
    try:
        import streamlit as st
        # This is a placeholder - actual detection would require JavaScript
        # For now, we'll show both options and let user choose
        return True
    except:
        return False


# ============================================================================
# DATA ENTRY TAB
# ============================================================================

def render_data_entry_tab():
    """Render Data Entry tab with professional QR scanning interface."""

    st.markdown("## Production Data Entry")
    st.caption("Scan battery pack QR code or enter ID manually")

    # Initialize session state
    if 'scanned_battery_id' not in st.session_state:
        st.session_state['scanned_battery_id'] = ''

    # QR Scanning Section
    st.markdown("---")
    st.markdown("### Battery Pack Identification")

    # Only show method selection if no scanner is currently open
    if not st.session_state.get('photo_upload_open', False) and not st.session_state.get('camera_scanner_open', False) and not st.session_state.get('barcode_scanner_open', False):
        # Three scanning methods - Barcode Scanner prioritized
        col1, col2, col3 = st.columns(3)

        with col1:
            st.markdown("""
            <div class="method-card primary">
                <div class="method-title">🔍 Barcode Scanner</div>
                <div class="method-description">
                    USB/Wireless Scanner<br/>
                    <strong>Recommended for Desktop</strong>
                </div>
            </div>
            """, unsafe_allow_html=True)

            if st.button("Use Barcode Scanner", key="open_barcode", use_container_width=True, type="primary"):
                st.session_state['barcode_scanner_open'] = True
                st.session_state['photo_upload_open'] = False
                st.session_state['camera_scanner_open'] = False
                st.rerun()

        with col2:
            st.markdown("""
            <div class="method-card secondary">
                <div class="method-title">📸 Photo Upload</div>
                <div class="method-description">
                    Upload QR image<br/>
                    Works on all devices
                </div>
            </div>
            """, unsafe_allow_html=True)

            if st.button("Upload QR Code Photo", key="open_upload", use_container_width=True):
                st.session_state['photo_upload_open'] = True
                st.session_state['camera_scanner_open'] = False
                st.session_state['barcode_scanner_open'] = False
                st.rerun()

        with col3:
            st.markdown("""
            <div class="method-card secondary">
                <div class="method-title">📷 Live Camera</div>
                <div class="method-description">
                    Direct scanning<br/>
                    Requires HTTPS
                </div>
            </div>
            """, unsafe_allow_html=True)

            if st.button("Open Camera Scanner", key="open_camera", use_container_width=True):
                st.session_state['camera_scanner_open'] = True
                st.session_state['photo_upload_open'] = False
                st.session_state['barcode_scanner_open'] = False
                st.rerun()

    # Barcode Scanner Input (Priority Method) - AUTO-SUBMIT ON ENTER
    if st.session_state.get('barcode_scanner_open', False):
        st.markdown("---")
        st.markdown("### Barcode Scanner Ready")

        st.markdown("""
        <div class="alert alert-success">
            <strong>✓ Scanner Active</strong><br/>
            Scan the QR code now - it will automatically proceed when done!
        </div>
        """, unsafe_allow_html=True)

        # Check if we have scanned data to process (set by JavaScript)
        if st.session_state.get('scanner_submit_trigger', False):
            scanner_data = st.session_state.get('scanner_input_auto', '')
            if scanner_data:
                # Extract battery ID from scanned data
                battery_id = extract_battery_id_from_url(scanner_data) if scanner_data.startswith('http') else scanner_data.strip()

                if battery_id:
                    st.session_state['scanned_battery_id'] = battery_id
                    st.session_state['barcode_scanner_open'] = False
                    st.session_state['scanner_input_auto'] = ''
                    st.session_state['scanner_submit_trigger'] = False
                    st.rerun()

        # Reset trigger flag
        st.session_state['scanner_submit_trigger'] = False

        # Text input field
        scanner_input = st.text_input(
            "Ready to Scan",
            key="scanner_input_auto",
            placeholder="Scan now - auto-submits on Enter...",
            help="Barcode scanner will auto-submit when it sends Enter key",
            label_visibility="visible"
        )

        # JavaScript for auto-focus and Enter key detection
        st.markdown("""
        <script>
            setTimeout(function() {
                const inputs = window.parent.document.querySelectorAll('input[placeholder*="auto-submits on Enter"]');
                if (inputs.length > 0) {
                    const inputField = inputs[0];

                    // Auto-focus immediately
                    inputField.focus();
                    inputField.select();

                    // Listen for Enter key (barcode scanners send Enter at the end)
                    inputField.addEventListener('keydown', function(event) {
                        if (event.key === 'Enter' || event.keyCode === 13) {
                            event.preventDefault();

                            // Blur to trigger Streamlit's state update
                            inputField.blur();

                            // Small delay then trigger rerun by pressing a hidden button
                            setTimeout(() => {
                                const buttons = window.parent.document.querySelectorAll('button[kind="secondary"]');
                                for (let btn of buttons) {
                                    if (btn.textContent.includes('Process Scan')) {
                                        btn.click();
                                        break;
                                    }
                                }
                            }, 50);
                        }
                    });

                    // Re-focus after blur (keep field focused)
                    inputField.addEventListener('blur', function() {
                        setTimeout(() => {
                            if (!document.activeElement || document.activeElement.tagName !== 'BUTTON') {
                                inputField.focus();
                            }
                        }, 100);
                    });
                }
            }, 100);
        </script>
        """, unsafe_allow_html=True)

        # Hidden button for JavaScript to trigger (sets flag and reruns)
        col1, col2, col3 = st.columns([2, 2, 1])
        with col2:
            if st.button("Process Scan", key="process_scan_trigger", use_container_width=True, type="secondary"):
                st.session_state['scanner_submit_trigger'] = True
                st.rerun()

        with col3:
            if st.button("Cancel", key="close_barcode", use_container_width=True):
                st.session_state['barcode_scanner_open'] = False
                st.session_state['scanner_input_auto'] = ''
                st.session_state['scanner_submit_trigger'] = False
                st.rerun()

    # Photo Upload Scanner
    if st.session_state.get('photo_upload_open', False):
        st.markdown("---")
        st.markdown("### Upload QR Code Image")

        st.markdown("""
        <div class="alert alert-info">
            <strong>Instructions:</strong> Upload a clear photo of the QR code.
            The battery ID will be detected and applied automatically.
        </div>
        """, unsafe_allow_html=True)

        uploaded_file = st.file_uploader(
            "Select QR Code Image",
            type=['png', 'jpg', 'jpeg', 'webp'],
            key="qr_upload",
            help="Upload a clear photo of the QR code"
        )

        if uploaded_file is not None:
            try:
                image = Image.open(uploaded_file)

                col_img, col_result = st.columns([1, 1])

                with col_img:
                    st.image(image, caption="Uploaded Image", use_column_width=True)

                with col_result:
                    with st.spinner("Scanning QR code..."):
                        qr_data = decode_qr_from_image(image)

                    if qr_data:
                        battery_id = extract_battery_id_from_url(qr_data)

                        st.markdown(f"""
                        <div class="alert alert-success">
                            <strong>QR Code Detected</strong><br/>
                            <span style="font-size: 1.75rem; font-weight: 700; display: block; margin: 0.5rem 0;">{battery_id}</span>
                            <span style="font-size: 0.875rem; opacity: 0.8;">Proceeding to data entry...</span>
                        </div>
                        """, unsafe_allow_html=True)

                        # Automatically set and proceed
                        st.session_state['scanned_battery_id'] = battery_id
                        st.session_state['photo_upload_open'] = False

                        # Add a brief delay for user to see confirmation
                        import time
                        time.sleep(1)
                        st.rerun()
                    else:
                        st.markdown("""
                        <div class="alert alert-error">
                            <strong>No QR Code Detected</strong><br/>
                            Please ensure the image is clear, well-lit, and the QR code is centered.
                        </div>
                        """, unsafe_allow_html=True)

            except Exception as e:
                st.error(f"Error processing image: {str(e)}")
                logger.error(f"Upload error: {e}", exc_info=True)

        if st.button("Cancel", key="close_upload"):
            st.session_state['photo_upload_open'] = False
            st.rerun()

    # Camera Scanner
    if st.session_state.get('camera_scanner_open', False):
        st.markdown("---")
        st.markdown("### Live Camera Scanner")

        st.markdown("""
        <div class="alert alert-info">
            Point your camera at the QR code. The battery ID will be detected automatically.
        </div>
        """, unsafe_allow_html=True)

        # Hidden input to receive scanned ID from JavaScript
        camera_scanned_id = st.text_input(
            "Scanned ID",
            key="camera_scanned_input",
            label_visibility="collapsed"
        )

        # Auto-submit when ID is detected
        if camera_scanned_id:
            st.session_state['scanned_battery_id'] = camera_scanned_id
            st.session_state['camera_scanner_open'] = False
            st.rerun()

        scanner_html = """
        <div class="scanner-container">
            <div id="reader" style="width:100%; max-width: 500px; margin: 0 auto;"></div>
            <div id="result" style="margin-top: 1rem; font-size: 1.1rem; font-weight: 600; text-align: center; color: #2E7D32;"></div>
        </div>

        <script src="https://unpkg.com/html5-qrcode@2.3.8/html5-qrcode.min.js"></script>
        <script>
            function onScanSuccess(decodedText, decodedResult) {
                let batteryId = decodedText;
                if (decodedText.includes('/entry/')) {
                    batteryId = decodedText.split('/entry/')[1].split('?')[0].split('#')[0];
                }

                document.getElementById('result').innerHTML = 'Detected: ' + batteryId;

                // Stop scanner
                html5QrcodeScanner.clear().catch(err => console.error(err));

                // Fill the hidden input field
                const inputElements = window.parent.document.querySelectorAll('input[aria-label="Scanned ID"]');
                if (inputElements.length > 0) {
                    const nativeInputValueSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, "value").set;
                    nativeInputValueSetter.call(inputElements[0], batteryId);
                    const event = new Event('input', { bubbles: true });
                    inputElements[0].dispatchEvent(event);
                }
            }

            function onScanFailure(error) {
                // Silent
            }

            let html5QrcodeScanner = new Html5QrcodeScanner(
                "reader",
                {
                    fps: 10,
                    qrbox: { width: 250, height: 250 },
                    aspectRatio: 1.0
                },
                false
            );

            html5QrcodeScanner.render(onScanSuccess, onScanFailure);
        </script>
        """

        st.components.v1.html(scanner_html, height=600)

        if st.button("Cancel", key="close_camera", use_container_width=True):
            st.session_state['camera_scanner_open'] = False
            st.rerun()

    # Display current battery ID
    battery_pack_id = st.session_state.get('scanned_battery_id', '')

    # Only show manual entry if no scanning interface is open and no ID is set
    if not battery_pack_id and not st.session_state.get('photo_upload_open', False) and not st.session_state.get('camera_scanner_open', False) and not st.session_state.get('barcode_scanner_open', False):
        st.markdown("---")
        with st.expander("Manual Entry"):
            st.caption("Enter Battery Pack ID manually if scanning is unavailable")
            manual_id = st.text_input(
                "Battery Pack ID",
                key="manual_battery_id",
                placeholder="Enter battery pack ID"
            )

            if st.button("Set ID", key="confirm_manual", use_container_width=True):
                if manual_id:
                    st.session_state['scanned_battery_id'] = manual_id
                    st.rerun()
                else:
                    st.error("Please enter a Battery Pack ID")

    if not battery_pack_id:
        return

    # Show selected battery ID with clear separator
    st.markdown("---")
    st.markdown(f"""
    <div class="alert alert-success">
        <strong>Current Battery Pack:</strong>
        <span style="font-size: 1.5rem; font-weight: 700; display: block; margin: 0.5rem 0;">{battery_pack_id}</span>
    </div>
    """, unsafe_allow_html=True)

    # Check if battery ID already exists
    exists_info = check_battery_exists(battery_pack_id)

    if exists_info['qr_exists'] or exists_info['data_exists']:
        st.warning(f"""
        **Battery Pack ID Already Exists**

        - QR Code: {'Exists' if exists_info['qr_exists'] else 'Not found'}
        - Production Data: {'Exists' if exists_info['data_exists'] else 'Not found'}

        This battery pack ID is already in the system. You are viewing/updating existing data.
        """)

    if st.button("Scan Different Pack", key="rescan", use_container_width=False):
        st.session_state['scanned_battery_id'] = ''
        st.session_state['photo_upload_open'] = False
        st.session_state['camera_scanner_open'] = False
        st.session_state['barcode_scanner_open'] = False
        st.rerun()

    st.markdown("---")

    # Process Selection
    st.markdown("### Process Selection")

    process_name = st.selectbox(
        "Select Production Process",
        options=list(PROCESS_DEFINITIONS.keys()),
        key="process_name"
    )

    # Check if this process already has data
    process_status = check_process_data_exists(battery_pack_id, process_name)

    # For standard processes (1-7), handle start/complete workflow
    if process_status['process_type'] == 'standard':
        if process_status['completed']:
            # Process fully completed - show edit option
            st.success(f"""
            **Process Completed ✓**

            Process "{process_name}" for battery pack {battery_pack_id} has been completed.
            """)
            st.markdown('<span class="badge badge-success">Completed</span>', unsafe_allow_html=True)

            st.markdown("---")

            # Show edit button
            col1, col2 = st.columns(2)
            with col1:
                if st.button("📝 Edit Data", use_container_width=True, type="secondary"):
                    st.session_state['edit_mode'] = True
                    st.rerun()

            # If not in edit mode, stop here
            if not st.session_state.get('edit_mode', False):
                return

            # In edit mode - show form below
            st.info("**Edit Mode:** Modify the data below and click Save to update.")

        elif process_status['has_any_data'] and process_status['both_modules_complete'] and not process_status['completed']:
            # Both modules filled, but not marked complete - show complete button + continue to form
            st.info(f"""
            **Both Modules Complete - Ready to Finalize**

            Both Module X and Module Y data have been entered for process "{process_name}".

            You can mark this process as complete, or continue editing the data below.
            """)
            st.markdown('<span class="badge badge-info">Ready to Complete</span>', unsafe_allow_html=True)

            st.markdown("---")

            # Show complete button
            col1, col2 = st.columns(2)
            with col1:
                if st.button("✓ Complete Process", type="primary", use_container_width=True):
                    try:
                        complete_process(battery_pack_id, process_name)
                        st.success(f"Process '{process_name}' completed successfully! End time recorded.")
                        st.balloons()
                        # Clear edit mode
                        if 'edit_mode' in st.session_state:
                            del st.session_state['edit_mode']
                        st.rerun()
                    except Exception as e:
                        st.error(f"Failed to complete process: {str(e)}")

            # Continue to show form for editing

        elif process_status['has_any_data'] and not process_status['both_modules_complete']:
            # Partial data - show what's missing
            missing_info = []
            if not process_status['module_x_complete']:
                missing_info.append("Module X")
            if not process_status['module_y_complete']:
                missing_info.append("Module Y")

            st.warning(f"""
            **Partial Data Entry**

            Some data has been entered, but the process is incomplete.

            Missing data for: **{', '.join(missing_info)}**

            Continue filling in the data below.
            """)
            st.markdown('<span class="badge badge-warning">Partial - Continue Entry</span>', unsafe_allow_html=True)

        else:
            # New process
            st.markdown('<span class="badge badge-success">New Record - Create Mode</span>', unsafe_allow_html=True)

    else:
        # For non-standard processes (pack, dispatch)
        if process_status['exists']:
            st.warning(f"""
            **Process Data Already Exists!**

            Data has already been entered for process "{process_name}" for battery pack {battery_pack_id}.

            You can edit the data below. Click Save to update.
            """)
            st.markdown('<span class="badge badge-warning">Edit Mode</span>', unsafe_allow_html=True)
        else:
            st.markdown('<span class="badge badge-success">New Record - Create Mode</span>', unsafe_allow_html=True)

    process_def = PROCESS_DEFINITIONS.get(process_name, {})
    qc_checks = process_def.get("qc_checks", [])

    # Load existing data if available
    existing_data = {}
    existing_technician = ""
    existing_qc = ""
    existing_remarks = ""

    if process_status['has_any_data']:
        try:
            from database import get_qc_checks
            checks = get_qc_checks(battery_pack_id, process_name)

            if checks:
                # Get operator info from first check (all have same values)
                existing_technician = checks[0].get('technician_name', '')
                existing_qc = checks[0].get('qc_name', '')
                existing_remarks = checks[0].get('remarks', '')

                # Build existing data dict
                for check in checks:
                    check_name = check.get('check_name', '')
                    existing_data[check_name] = {
                        'module_x': check.get('module_x', ''),
                        'module_y': check.get('module_y', '')
                    }
        except Exception as e:
            logger.error(f"Error loading existing data: {e}")

    st.markdown("---")

    # Data Entry Form
    st.markdown(f"### Process: {process_name}")
    st.caption(process_def.get('work_description', ''))

    # Operator Information
    col_tech, col_qc = st.columns(2)

    with col_tech:
        technician_name = st.text_input(
            "Technician Name",
            value=existing_technician,
            key="technician_name",
            placeholder="Enter technician name"
        )

    with col_qc:
        qc_name = st.text_input(
            "QC Inspector (Optional)",
            value=existing_qc,
            key="qc_name",
            placeholder="Enter QC inspector name"
        )

    remarks = st.text_area(
        "Remarks (Optional)",
        value=existing_remarks,
        key="remarks",
        placeholder="Additional notes or observations",
        height=100
    )

    st.markdown("---")

    # Quality Control Checks
    st.markdown("### Quality Control Checks")

    if not qc_checks:
        st.warning("No QC checks defined for this process")
        return

    checks_data = {}

    for idx, check_name in enumerate(qc_checks):
        st.markdown(f"""
        <div class="card">
            <div class="card-header">Check {idx+1}: {check_name}</div>
        </div>
        """, unsafe_allow_html=True)

        # Get existing values for this check
        existing_check = existing_data.get(check_name, {})
        existing_module_x = existing_check.get('module_x', '')
        existing_module_y = existing_check.get('module_y', '')

        # Find index of existing value in options
        options = ["", "OK", "NOT OK", "N/A"]
        default_x_index = options.index(existing_module_x) if existing_module_x in options else 0
        default_y_index = options.index(existing_module_y) if existing_module_y in options else 0

        col_x, col_y = st.columns(2)

        with col_x:
            module_x = st.radio(
                "Module X",
                options=options,
                index=default_x_index,
                key=f"check_{idx}_{check_name.replace(' ', '_').replace('/', '_')}_x",
                horizontal=True
            )

        with col_y:
            module_y = st.radio(
                "Module Y",
                options=options,
                index=default_y_index,
                key=f"check_{idx}_{check_name.replace(' ', '_').replace('/', '_')}_y",
                horizontal=True
            )

        if module_x or module_y:
            checks_data[check_name] = {
                "module_x": module_x if module_x else "",
                "module_y": module_y if module_y else ""
            }

    st.markdown("---")

    # Submit Button
    if st.button("Save Production Data", type="primary", use_container_width=True, key="save_data"):
        # Simple validation
        if not technician_name or not technician_name.strip():
            st.error("Technician name is required")
            return

        if len(technician_name) > 100:
            st.error("Technician name too long (max 100 characters)")
            return

        if qc_name and len(qc_name) > 100:
            st.error("QC name too long (max 100 characters)")
            return

        if remarks and len(remarks) > 500:
            st.error("Remarks too long (max 500 characters)")
            return

        if not checks_data:
            st.error("Please complete at least one QC check")
            return

        try:
            with st.spinner("Saving data..."):
                # Convert checks_data to list format
                checks_list = []
                for check_name, check_values in checks_data.items():
                    checks_list.append({
                        'check_name': check_name,
                        'module_x': check_values['module_x'],
                        'module_y': check_values['module_y']
                    })

                output_file = add_detailed_entry(
                    battery_pack_id=battery_pack_id,
                    process_name=process_name,
                    technician_name=technician_name,
                    qc_name=qc_name,
                    remarks=remarks,
                    checks=checks_list
                )

                st.success("Production data saved successfully")
                st.info(f"File: {output_file}")
                logger.info(f"Data saved for {battery_pack_id}, process: {process_name}")

                # Clear form and edit mode
                st.session_state['scanned_battery_id'] = ''
                st.session_state['photo_upload_open'] = False
                st.session_state['camera_scanner_open'] = False
                st.session_state['barcode_scanner_open'] = False
                if 'edit_mode' in st.session_state:
                    del st.session_state['edit_mode']
                for key in list(st.session_state.keys()):
                    if key.startswith('check_'):
                        del st.session_state[key]

                if st.button("Enter Next Pack", key="scan_next"):
                    st.rerun()

        except Exception as e:
            st.error(f"Failed to save: {str(e)}")
            logger.error(f"Data entry failed: {e}")


# ============================================================================
# QR CODE GENERATOR TAB
# ============================================================================

def render_qr_generator_tab():
    """Render QR Code Generator tab."""
    st.markdown("## QR Code Generator")
    st.caption("Generate and download QR codes for battery pack identification")

    # Add tabs for Generate and View Saved
    tab1, tab2 = st.tabs(["Generate New QR Code", "Saved QR Codes"])

    with tab1:
        st.markdown("### Create New QR Code")

        battery_pack_id = st.text_input(
            "Battery Pack ID",
            placeholder="Enter battery pack identifier",
            key="qr_battery_id"
        )

        col_a, col_b = st.columns(2)
        with col_a:
            size = st.selectbox("QR Code Size (px)", options=[200, 300, 400, 500], index=1, key="qr_size")
        with col_b:
            include_label = st.checkbox("Include Text Label", value=True, key="qr_label")

        if st.button("Generate QR Code", type="primary", use_container_width=True):
            if not battery_pack_id:
                st.error("Please enter a Battery Pack ID")
            else:
                # Check if battery ID already exists
                exists_info = check_battery_exists(battery_pack_id)

                if exists_info['qr_exists'] or exists_info['data_exists']:
                    st.warning(f"""
                    **Battery Pack ID Already Exists**

                    - QR Code exists: {'Yes' if exists_info['qr_exists'] else 'No'}
                    - Data exists: {'Yes' if exists_info['data_exists'] else 'No'}

                    This battery pack ID is already in use. Each battery must have a unique ID.
                    Please use a different ID or check existing records.
                    """)
                else:
                    try:
                        qr_bytes = generate_qr_code(battery_pack_id, size=size, include_label=include_label)
                        st.session_state['qr_image'] = qr_bytes
                        st.session_state['qr_pack_id'] = battery_pack_id
                        st.success(f"QR code generated and saved for {battery_pack_id}")
                        st.info(f"Saved to: qr_codes/{battery_pack_id}.png")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Generation failed: {str(e)}")

        if 'qr_image' in st.session_state:
            st.markdown("---")
            st.markdown("### Generated QR Code")
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                st.image(st.session_state['qr_image'], use_column_width=True)
            st.download_button(
                label="Download QR Code",
                data=st.session_state['qr_image'],
                file_name=f"{st.session_state['qr_pack_id']}.png",
                mime="image/png",
                use_container_width=True,
                type="primary"
            )

    with tab2:
        st.markdown("### Saved QR Codes")

        try:
            from pathlib import Path
            import os

            qr_dir = Path("qr_codes")
            if not qr_dir.exists():
                st.info("No QR codes generated yet. Use the 'Generate New QR Code' tab to create one.")
            else:
                # Get all QR code files
                qr_files = sorted(qr_dir.glob("*.png"), key=os.path.getmtime, reverse=True)

                if not qr_files:
                    st.info("No QR codes generated yet. Use the 'Generate New QR Code' tab to create one.")
                else:
                    st.markdown("---")

                    # Search and Filter (same format as Reports tab)
                    col_search, col_sort = st.columns([3, 1])

                    with col_search:
                        search_term = st.text_input(
                            "Search QR Codes",
                            placeholder="Enter battery pack ID to search",
                            key="qr_search"
                        )

                    with col_sort:
                        sort_order = st.selectbox(
                            "Sort By",
                            options=["Newest First", "Oldest First", "Name A-Z", "Name Z-A"],
                            key="sort_qr"
                        )

                    # Filter files based on search
                    if search_term:
                        filtered_files = [f for f in qr_files if search_term.lower() in f.stem.lower()]
                    else:
                        filtered_files = qr_files

                    # Sort files
                    if sort_order == "Name A-Z":
                        filtered_files = sorted(filtered_files, key=lambda x: x.stem)
                    elif sort_order == "Name Z-A":
                        filtered_files = sorted(filtered_files, key=lambda x: x.stem, reverse=True)
                    elif sort_order == "Oldest First":
                        filtered_files = sorted(filtered_files, key=os.path.getmtime)
                    else:  # Newest First
                        filtered_files = sorted(filtered_files, key=os.path.getmtime, reverse=True)

                    st.caption(f"Showing {len(filtered_files)} of {len(qr_files)} QR codes")

                    st.markdown("---")

                    # Display QR codes in list format (same as Reports tab)
                    for qr_file in filtered_files:
                        pack_id = qr_file.stem
                        col1, col2 = st.columns([4, 1])

                        with col1:
                            st.markdown(f"**{pack_id}**")
                            file_size = qr_file.stat().st_size / 1024
                            st.caption(f"Size: {file_size:.1f} KB | File: qr_codes/{pack_id}.png")

                        with col2:
                            # Read file for download
                            with open(qr_file, 'rb') as f:
                                qr_data = f.read()

                            st.download_button(
                                label="Download File",
                                data=qr_data,
                                file_name=f"{pack_id}.png",
                                mime="image/png",
                                use_container_width=True,
                                key=f"download_qr_{pack_id}"
                            )

                        st.markdown('<hr style="margin: 0.5rem 0;">', unsafe_allow_html=True)

                    st.markdown("---")

                    # Bulk Actions (same format as Reports tab)
                    st.markdown("### Bulk Actions")

                    if st.button("Download All QR Codes as ZIP", use_container_width=True):
                        try:
                            import zipfile
                            from io import BytesIO
                            from datetime import datetime

                            # Create ZIP file in memory
                            zip_buffer = BytesIO()
                            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                                for qr_file in filtered_files:
                                    zip_file.write(qr_file, qr_file.name)

                            zip_buffer.seek(0)
                            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

                            st.download_button(
                                label=f"Download ZIP File ({len(filtered_files)} files)",
                                data=zip_buffer.getvalue(),
                                file_name=f"QR_Codes_{timestamp}.zip",
                                mime="application/zip",
                                use_container_width=True,
                                type="primary"
                            )
                        except Exception as e:
                            st.error(f"Failed to create ZIP: {str(e)}")

        except Exception as e:
            st.error(f"Error loading QR codes: {str(e)}")
            logger.error(f"QR gallery error: {e}")


# ============================================================================
# DASHBOARD TAB
# ============================================================================

def render_dashboard_tab():
    """Render production dashboard with Pack Tracker and Production Charts."""
    st.markdown("## Production Dashboard")
    st.caption("Battery Pack Tracker and Production Analytics")

    # Add refresh button
    col_title, col_refresh = st.columns([6, 1])
    with col_refresh:
        if st.button("Refresh", key="refresh_dashboard"):
            st.rerun()

    try:
        # Read data directly from DATABASE for real-time accuracy
        st.markdown("---")

        # ZMC Pack Tracker Table
        st.markdown("### Battery Pack Tracker")

        tracker_data = []
        process_stages = ["Cell sorting", "Module assembly", "Pre Encapsulation", "Wire Bonding",
                         "Post Encapsulation", "EOL Testing", "Pack assembly", "Ready for Dispatch"]

        # Get all battery packs from database
        all_packs = get_all_battery_packs()

        if not all_packs:
            st.info("No production data available. Begin tracking battery packs to see metrics here.")
            return

        for idx, pack_id in enumerate(sorted(all_packs), 1):
            try:
                row_data = {"Sl.No": idx, "Battery Pack": pack_id}

                # Get all QC checks for this pack from database
                all_checks = get_qc_checks(pack_id)

                # Group checks by process
                processes_completed = {}
                for check in all_checks:
                    process_name = check['process_name']
                    if process_name not in processes_completed:
                        processes_completed[process_name] = []

                    # Check if any result is NOT OK
                    module_x = check.get('module_x', '')
                    module_y = check.get('module_y', '')

                    if "NOT OK" in str(module_x) or "NOT OK" in str(module_y):
                        processes_completed[process_name].append("NOT OK")
                    elif "OK" in str(module_x) or "OK" in str(module_y):
                        processes_completed[process_name].append("OK")

                # Map database processes to display stages
                qc_ok_count = 0
                processes_with_data = 0  # Count all processes with any data
                has_deviations = False

                for stage in process_stages:
                    stage_status = "0"

                    # Check if this stage has data in database
                    if stage in processes_completed:
                        checks = processes_completed[stage]
                        if len(checks) > 0:
                            processes_with_data += 1
                            if "NOT OK" in checks:
                                stage_status = "OK with Deviation"
                                has_deviations = True
                            else:
                                stage_status = "QC OK"
                                qc_ok_count += 1

                    row_data[stage] = stage_status

                # Determine final status based on processes completed
                total_processes_done = processes_with_data
                if total_processes_done >= 8 and not has_deviations:
                    row_data["Status"] = "Ready to dispatch"
                elif total_processes_done > 0:
                    row_data["Status"] = "In Process"
                else:
                    row_data["Status"] = "Not Started"

                tracker_data.append(row_data)

            except Exception as e:
                logger.error(f"Error reading pack {pack_id}: {e}")
                continue

        # Create DataFrame
        if tracker_data:
            df_tracker = pd.DataFrame(tracker_data)

            # Style the dataframe
            def style_cell(val):
                if val == "QC OK":
                    return 'background-color: #c8e6c9; color: #2e7d32; font-weight: 500;'
                elif val == "OK with Deviation":
                    return 'background-color: #fff9c4; color: #f57c00; font-weight: 500;'
                elif val == "Ready to dispatch":
                    return 'background-color: #2e7d32; color: white; font-weight: 700; text-align: center;'
                elif val == "In Process":
                    return 'background-color: #1976d2; color: white; font-weight: 600; text-align: center;'
                elif val == "Not Started":
                    return 'background-color: #9e9e9e; color: white; font-weight: 500; text-align: center;'
                elif val == "0":
                    return 'background-color: #f5f5f5; color: #9e9e9e;'
                return ''

            # Display as styled table
            st.dataframe(
                df_tracker.style.applymap(style_cell),
                use_container_width=True,
                height=400
            )

        st.markdown("---")

        # Production Charts
        col_chart1, col_chart2 = st.columns(2)

        # Calculate metrics
        total_packs = len(all_packs)
        target_packs = 50  # This could be configured

        # Count packs by status
        completed_packs = sum(1 for row in tracker_data if row.get("Status") == "Ready to dispatch")
        in_process_packs = sum(1 for row in tracker_data if row.get("Status") == "In Process")
        not_started_packs = sum(1 for row in tracker_data if row.get("Status") == "Not Started")
        rejected_packs = 0  # Could track packs with critical failures

        with col_chart1:
            st.markdown("### Plan vs Actual in Packs")

            # Bar chart data
            bar_data = pd.DataFrame({
                'Category': ['Target', 'Completed', 'In Process', 'Rejected'],
                'Count': [target_packs, completed_packs, in_process_packs, rejected_packs]
            })

            # Create plotly bar chart
            fig_bar = go.Figure(data=[
                go.Bar(
                    x=bar_data['Category'],
                    y=bar_data['Count'],
                    marker_color=['#1976D2', '#FF9800', '#FFA726', '#EF5350'],
                    text=bar_data['Count'],
                    textposition='outside',
                    textfont=dict(size=14, color='black', family='Arial Black')
                )
            ])

            fig_bar.update_layout(
                height=350,
                margin=dict(l=20, r=20, t=40, b=40),
                yaxis=dict(range=[0, max(target_packs, completed_packs, in_process_packs) * 1.2]),
                plot_bgcolor='white',
                paper_bgcolor='white',
                font=dict(size=12, color='#424242')
            )

            st.plotly_chart(fig_bar, use_container_width=True)

        with col_chart2:
            st.markdown("### Production Plan in %")

            # Calculate percentages
            total_actual = completed_packs + in_process_packs + rejected_packs
            if target_packs > 0:
                completed_pct = (completed_packs / target_packs) * 100
                in_process_pct = (in_process_packs / target_packs) * 100
                rejected_pct = (rejected_packs / target_packs) * 100
                target_pct = 100
            else:
                completed_pct = in_process_pct = rejected_pct = target_pct = 0

            # Pie chart data
            pie_data = pd.DataFrame({
                'Status': ['Target', 'Completed', 'In Process', 'Rejected'],
                'Percentage': [target_pct, completed_pct, in_process_pct, rejected_pct]
            })

            # Create plotly pie chart
            fig_pie = go.Figure(data=[go.Pie(
                labels=pie_data['Status'],
                values=pie_data['Percentage'],
                marker=dict(colors=['#1976D2', '#FF9800', '#4CAF50', '#EF5350']),
                textinfo='label+percent',
                textfont=dict(size=12, color='white', family='Arial'),
                hole=0
            )])

            fig_pie.update_layout(
                height=350,
                margin=dict(l=20, r=20, t=20, b=20),
                showlegend=True,
                legend=dict(
                    orientation="v",
                    yanchor="middle",
                    y=0.5,
                    xanchor="left",
                    x=1.05
                ),
                plot_bgcolor='white',
                paper_bgcolor='white'
            )

            st.plotly_chart(fig_pie, use_container_width=True)

    except Exception as e:
        st.error(f"Error loading dashboard: {str(e)}")
        logger.error(f"Dashboard error: {e}")


# ============================================================================
# REPORTS TAB
# ============================================================================

def render_reports_tab():
    """Render reports management interface."""
    st.markdown("## Production Reports")
    st.caption("View, search, and download production reports")

    try:
        # Read from master sample.xlsx file
        master_file = Path("sample.xlsx")

        if not master_file.exists():
            st.info("No reports found. sample.xlsx not found.")
            return

        # Load workbook to get sheet names
        wb = openpyxl.load_workbook(master_file, read_only=True)
        battery_sheets = [sheet for sheet in wb.sheetnames[1:]]  # Skip template
        wb.close()

        if not battery_sheets:
            st.info("No reports available. Generate QR codes and enter production data to create reports.")
            return

        st.markdown("---")

        # Search and Filter
        col_search, col_sort = st.columns([3, 1])

        with col_search:
            search_term = st.text_input(
                "Search Reports",
                placeholder="Enter battery pack ID to search",
                key="search_reports"
            )

        with col_sort:
            sort_order = st.selectbox(
                "Sort By",
                options=["Newest First", "Oldest First", "Name A-Z", "Name Z-A"],
                key="sort_reports"
            )

        # Filter sheets
        if search_term:
            filtered_sheets = [s for s in battery_sheets if search_term.lower() in s.lower()]
        else:
            filtered_sheets = battery_sheets

        # Sort sheets
        if sort_order == "Name A-Z":
            filtered_sheets = sorted(filtered_sheets)
        elif sort_order == "Name Z-A":
            filtered_sheets = sorted(filtered_sheets, reverse=True)
        else:
            # For time-based sorting, just use alphabetical
            filtered_sheets = sorted(filtered_sheets)

        st.caption(f"Showing {len(filtered_sheets)} of {len(battery_sheets)} reports")

        st.markdown("---")

        # Display sheets
        for sheet_name in filtered_sheets:
            col1, col2 = st.columns([4, 1])

            with col1:
                st.markdown(f"**{sheet_name}**")
                st.caption(f"Sheet in: sample.xlsx")

            with col2:
                # Download individual battery pack Excel file
                individual_file = Path("excel_reports") / f"{sheet_name}.xlsx"

                # Generate if doesn't exist
                if not individual_file.exists():
                    try:
                        generate_battery_excel(sheet_name)
                    except Exception as e:
                        logger.error(f"Error generating Excel for {sheet_name}: {e}")

                # Now try to download
                if individual_file.exists():
                    with open(individual_file, 'rb') as f:
                        file_data = f.read()
                    st.download_button(
                        label="Download File",
                        data=file_data,
                        file_name=f"{sheet_name}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"download_{sheet_name}",
                        use_container_width=True
                    )
                else:
                    st.caption("File not found")

            st.markdown('<hr style="margin: 0.5rem 0;">', unsafe_allow_html=True)

        st.markdown("---")

        # Bulk Export
        st.markdown("### Bulk Actions")

        # Download Complete File button
        with open(master_file, 'rb') as f:
            master_data = f.read()
        st.download_button(
            label="Download Complete File (sample.xlsx)",
            data=master_data,
            file_name="sample.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_complete",
            use_container_width=True
        )

        # Generate CSV data from database
        try:
            import io
            csv_buffer = io.StringIO()

            # Get all battery packs and their data
            all_packs = get_all_battery_packs()

            # Write CSV header
            csv_buffer.write("Battery Pack ID,Process Name,Check Name,Module X,Module Y,Technician,QC Name,Remarks,Start Date,End Date\n")

            # Write data rows
            for pack_id in all_packs:
                checks = get_qc_checks(pack_id)
                for check in checks:
                    csv_buffer.write(f"{pack_id},")
                    csv_buffer.write(f"{check.get('process_name', '')},")
                    csv_buffer.write(f"{check.get('check_name', '')},")
                    csv_buffer.write(f"{check.get('module_x', '')},")
                    csv_buffer.write(f"{check.get('module_y', '')},")
                    csv_buffer.write(f"{check.get('technician_name', '')},")
                    csv_buffer.write(f"{check.get('qc_name', '')},")
                    csv_buffer.write(f"{check.get('remarks', '')},")
                    csv_buffer.write(f"{check.get('start_date', '')},")
                    csv_buffer.write(f"{check.get('end_date', '')}\n")

            csv_data = csv_buffer.getvalue()

            st.download_button(
                label="Download CSV Report",
                data=csv_data,
                file_name=f"production_reports_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mime="text/csv",
                use_container_width=True
            )
        except Exception as e:
            logger.error(f"Error generating CSV: {e}")

        # Database Backup Section
        st.markdown("---")
        st.markdown("### Database Backup & Recovery")

        col_info, col_backup = st.columns([2, 1])

        with col_info:
            db_size = get_database_size()
            st.info(f"Current Database Size: **{db_size} MB**")

            backups = list_backups()
            if backups:
                latest_backup = backups[0]
                st.success(f"Latest Backup: {latest_backup['filename']} ({latest_backup['size_mb']} MB, {latest_backup['age_days']} days ago)")
                st.caption(f"Total Backups: {len(backups)}")
            else:
                st.warning("No backups found")

        with col_backup:
            if st.button("Create Backup Now", type="primary", use_container_width=True):
                with st.spinner("Creating backup..."):
                    backup_file = create_backup()
                    if backup_file:
                        st.success(f"Backup created: {backup_file.name}")
                        st.rerun()
                    else:
                        st.error("Backup failed")

        # Show backup history
        if backups:
            st.markdown("#### Backup History")
            backup_data = []
            for backup in backups[:10]:  # Show last 10 backups
                backup_data.append({
                    "Filename": backup['filename'],
                    "Size (MB)": backup['size_mb'],
                    "Created": backup['created'].strftime("%Y-%m-%d %H:%M:%S"),
                    "Age (days)": backup['age_days']
                })

            import pandas as pd
            df_backups = pd.DataFrame(backup_data)
            st.dataframe(df_backups, use_container_width=True, hide_index=True)

    except Exception as e:
        st.error(f"Error loading reports: {str(e)}")
        logger.error(f"Reports error: {e}")


# ============================================================================
# MAIN APPLICATION
# ============================================================================

def main():
    """Main application entry point."""

    # Header
    st.title("Battery Pack MES")
    st.caption("Manufacturing Execution System")

    # Navigation Tabs
    tab1, tab2, tab3, tab4 = st.tabs([
        "Data Entry",
        "QR Generator",
        "Dashboard",
        "Reports"
    ])

    with tab1:
        render_data_entry_tab()

    with tab2:
        render_qr_generator_tab()

    with tab3:
        render_dashboard_tab()

    with tab4:
        render_reports_tab()


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        logger.error(f"Application error: {e}", exc_info=True)
        st.error(f"Application error: {str(e)}")

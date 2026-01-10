#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Complete cleanup script for Battery Pack MES
Deletes ALL test data including database records and generated files

SAFETY FEATURES:
- Preserves database structure
- Keeps sample.xlsx template intact
- Confirmation prompt before deletion
- Graceful error handling
"""

import sqlite3
import os
from pathlib import Path
import sys
import openpyxl
import shutil

def main():
    print("=" * 70)
    print("  BATTERY PACK MES - DATA CLEANUP UTILITY")
    print("=" * 70)

    # Get paths
    script_dir = Path(__file__).parent
    db_path = script_dir / "battery_mes.db"
    qr_codes_dir = script_dir / "qr_codes"
    excel_reports_dir = script_dir / "excel_reports"
    sample_file = script_dir / "sample.xlsx"

    # Check if database exists
    if not db_path.exists():
        print(f"\n[!] Database not found at: {db_path}")
        print("Nothing to clean up.")
        return

    # Connect to database
    try:
        conn = sqlite3.connect(str(db_path))
        cursor = conn.cursor()
    except Exception as e:
        print(f"\n[ERROR] Cannot connect to database: {str(e)}")
        return

    # Count current records
    try:
        cursor.execute("SELECT COUNT(*) FROM battery_packs")
        pack_count = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM qc_checks")
        check_count = cursor.fetchone()[0]
    except Exception as e:
        print(f"\n[ERROR] Error reading database: {str(e)}")
        conn.close()
        return

    # Count files (excluding .gitkeep and sample.xlsx)
    qr_files = []
    if qr_codes_dir.exists():
        qr_files = [f for f in qr_codes_dir.glob("*.png") if f.name != ".gitkeep"]

    excel_files = []
    if excel_reports_dir.exists():
        excel_files = [f for f in excel_reports_dir.glob("*.xlsx")
                      if f.name != ".gitkeep"]

    # Display what will be deleted
    print(f"\nCURRENT DATA STATUS:\n")
    print(f"  Database Records:")
    print(f"    - Battery Packs: {pack_count} records")
    print(f"    - QC Checks: {check_count} records")
    print(f"\n  Generated Files:")
    print(f"    - QR Codes: {len(qr_files)} PNG files")
    print(f"    - Excel Reports: {len(excel_files)} files")

    print(f"\n[CLEANED] sample.xlsx:")
    print(f"    - Battery pack sheets will be removed")
    print(f"    - Template sheet will be preserved")
    print(f"\n[PRESERVED] The following will NOT be deleted:")
    print(f"    - Database structure (tables, schema)")
    print(f"    - sample.xlsx template structure")
    print(f"    - All application code")
    print(f"    - All configuration files")

    if pack_count == 0 and check_count == 0 and len(qr_files) == 0 and len(excel_files) == 0:
        print(f"\n[OK] System is already clean. Nothing to delete.")
        conn.close()
        return

    print(f"\n[WARNING] Data will be PERMANENTLY DELETED!")
    print("=" * 70)

    # Confirmation prompt
    response = input("\nType 'DELETE ALL' to confirm (or press Enter to cancel): ")

    if response.strip() != "DELETE ALL":
        print("\n[CANCELLED] No data was deleted.")
        conn.close()
        return

    print("\n[PROCESSING] Starting cleanup...\n")

    deleted_items = {
        'packs': 0,
        'checks': 0,
        'qr_files': 0,
        'excel_files': 0
    }

    try:
        # Enable foreign keys (needed for cascade to work in SQLite)
        cursor.execute("PRAGMA foreign_keys = ON")

        # Delete from database - delete QC checks first, then battery packs
        if check_count > 0:
            cursor.execute("DELETE FROM qc_checks")
            print(f"  [OK] Deleted {check_count} QC check records")
            deleted_items['checks'] = check_count

        if pack_count > 0:
            cursor.execute("DELETE FROM battery_packs")
            print(f"  [OK] Deleted {pack_count} battery pack records")
            deleted_items['packs'] = pack_count

        conn.commit()

        # Delete QR code files (preserve .gitkeep)
        for qr_file in qr_files:
            try:
                qr_file.unlink()
                deleted_items['qr_files'] += 1
            except Exception as e:
                print(f"  [WARNING] Could not delete {qr_file.name}: {str(e)}")

        if deleted_items['qr_files'] > 0:
            print(f"  [OK] Deleted {deleted_items['qr_files']} QR code files")

        # Delete Excel report files (preserve .gitkeep, DON'T touch sample.xlsx)
        for excel_file in excel_files:
            try:
                # Extra safety check - never delete sample.xlsx
                if excel_file.name == "sample.xlsx":
                    continue
                excel_file.unlink()
                deleted_items['excel_files'] += 1
            except Exception as e:
                print(f"  [WARNING] Could not delete {excel_file.name}: {str(e)}")

        if deleted_items['excel_files'] > 0:
            print(f"  [OK] Deleted {deleted_items['excel_files']} Excel report files")

        # Clean sample.xlsx (remove all battery pack sheets, keep template structure)
        if sample_file.exists():
            try:
                # Backup sample.xlsx first
                backup_file = script_dir / "sample.xlsx.backup"
                shutil.copy2(sample_file, backup_file)

                wb = openpyxl.load_workbook(sample_file)
                original_sheets = len(wb.sheetnames)
                sheets_removed = 0

                # Remove all sheets except the first one (template)
                # Keep the first sheet as template
                sheets_to_remove = wb.sheetnames[1:]  # All sheets except first

                for sheet_name in sheets_to_remove:
                    try:
                        wb.remove(wb[sheet_name])
                        sheets_removed += 1
                    except:
                        pass

                # Save cleaned sample.xlsx
                wb.save(sample_file)
                wb.close()

                if sheets_removed > 0:
                    print(f"  [OK] Cleaned sample.xlsx: removed {sheets_removed} battery pack sheets")
                    print(f"       (Template structure preserved)")

                # Remove backup if everything went well
                if backup_file.exists():
                    backup_file.unlink()

            except Exception as e:
                print(f"  [WARNING] Could not clean sample.xlsx: {str(e)}")
                # Restore from backup if error occurred
                if backup_file.exists():
                    shutil.copy2(backup_file, sample_file)
                    backup_file.unlink()
                    print(f"  [OK] Restored sample.xlsx from backup")

        # Verify database cleanup
        cursor.execute("SELECT COUNT(*) FROM battery_packs")
        remaining_packs = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM qc_checks")
        remaining_checks = cursor.fetchone()[0]

        print("\n" + "=" * 70)
        print("[SUCCESS] CLEANUP COMPLETE!")
        print("=" * 70)

        print(f"\nFinal Status:")
        print(f"  Database:")
        print(f"    - Battery Packs: {remaining_packs} records")
        print(f"    - QC Checks: {remaining_checks} records")
        print(f"\n  Files Remaining:")
        print(f"    - sample.xlsx: PRESERVED (template intact)")
        print(f"    - Database structure: INTACT")

        if remaining_packs == 0 and remaining_checks == 0:
            print(f"\n[SUCCESS] System reset to clean state.")
            print(f"          You can now start with fresh production data.")
        else:
            print(f"\n[WARNING] Some records still remain in database.")
            print(f"          This may indicate a problem with the cleanup.")

    except Exception as e:
        print(f"\n[ERROR] Error during cleanup: {str(e)}")
        print(f"        Rolling back database changes...")
        conn.rollback()
        print(f"\n        Database has been restored to previous state.")
        print(f"        Some files may have been deleted before error occurred.")
        conn.close()
        sys.exit(1)

    finally:
        conn.close()

    print("\n" + "=" * 70)

if __name__ == "__main__":
    main()
